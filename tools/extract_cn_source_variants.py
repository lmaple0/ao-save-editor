from __future__ import annotations

from collections import OrderedDict
from pathlib import Path
import json

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = Path.home() / "Downloads" / "ao_data_cn.xlsx"
CANONICAL = ROOT / "ao_zh_glossary.json"
OUT_WORKSPACE = ROOT / "ao_cn_source_variants_v4.json"
OUT_PUBLISH = ROOT.parent / "publish" / "package" / "ao_cn_source_variants_v4.json"


CORE_ALIAS = OrderedDict([
    ("护盾", "盾牌"),
    ("神盾", "神佑"),
    ("小丑", "小丑"),
    ("圣典", "法则"),
    ("权杖", "王权"),
    ("智者", "贤者"),
    ("力量", "力量"),
    ("战斧", "战斧"),
    ("骑兵", "骑士"),
    ("妖精", "妖精"),
    ("羽翼", "羽翼"),
    ("草薙", "草薤"),
    ("利爪", "利爪"),
    ("坚守", "守护"),
    ("罪恶", "罪恶"),
    ("白金", "白金"),
    ("永恒", "无限"),
    ("纹章", "纹章"),
    ("幻象", "幻象"),
    ("魅影", "魅影"),
    ("灵猫", "灵猫"),
])


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def repair_core_text(value):
    if not isinstance(value, str):
        return value
    # Only repair obviously broken core-quartz labels.
    if any("\uac00" <= ch <= "\ud7af" for ch in value) or sum(1 for ch in value if "\u4e00" <= ch <= "\u9fff") == 0:
        candidates = [value]
        for encoding in ("cp949", "cp950"):
            try:
                candidates.append(value.encode(encoding).decode("cp936"))
            except Exception:
                pass
        return max(candidates, key=lambda s: sum(1 for ch in s if "\u4e00" <= ch <= "\u9fff"))
    return value


def as_variant(row: dict | None) -> dict | None:
    if not row:
        return None
    return {
        "id_hex": row.get("id_hex"),
        "id_dec": row.get("id_dec"),
        "key": row.get("key"),
        "reference_no": row.get("reference_no"),
        "owner_key": row.get("owner_key"),
        "owner_zh": row.get("owner_zh"),
        "zh_joyoland": row.get("zh_cn"),
        "category_zh": row.get("category_zh"),
        "skill_kind": row.get("skill_kind"),
        "source": row.get("source"),
    }


def index_items(items: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for row in items:
        name = row.get("zh_cn")
        if not name:
            continue
        out.setdefault(name, []).append(row)
    return out


def chunk_rows(ws, start_row: int = 2):
    for r in range(start_row, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        yield r


def build_characters(ws, glossary: dict) -> dict:
    chars = {row["zh_cn"]: row for row in glossary.get("characters", [])}
    formulas = {row["zh_cn"]: row for row in glossary.get("level_formulas", [])}
    rows = []
    for r in chunk_rows(ws):
        source_name = ws.cell(r, 1).value
        canonical = chars.get(source_name)
        formula = formulas.get(source_name)
        rows.append({
            "row": r,
            "source_name": source_name,
            "zh_cle": source_name,
            "zh_joyoland": canonical["zh_cn"] if canonical else source_name,
            "canonical_character": {
                "key": canonical.get("key") if canonical else None,
                "zh_joyoland": canonical.get("zh_cn") if canonical else None,
                "base_offset_hex": canonical.get("base_offset_hex") if canonical else None,
                "source": canonical.get("source") if canonical else None,
            } if canonical else None,
            "canonical_level_formula": formula,
            "source_stats": {
                "lv": ws.cell(r, 2).value,
                "hp": ws.cell(r, 3).value,
                "ep": ws.cell(r, 4).value,
                "str": ws.cell(r, 5).value,
                "def": ws.cell(r, 6).value,
                "ats": ws.cell(r, 7).value,
                "adf": ws.cell(r, 8).value,
                "spd": ws.cell(r, 9).value,
                "mov": ws.cell(r, 10).value,
                "dex": ws.cell(r, 11).value,
                "agl": ws.cell(r, 12).value,
                "rng": ws.cell(r, 13).value,
            },
            "match_status": "exact_name" if canonical else "unmatched",
        })
    return {
        "sheet": ws.title,
        "kind": "character_source_variants",
        "row_count": len(rows),
        "rows": rows,
    }


def build_reference_skills(ws, glossary: dict) -> dict:
    ref = glossary.get("reference_skills", [])
    rows = []
    for idx, r in enumerate(chunk_rows(ws)):
        canonical = ref[idx] if idx < len(ref) else None
        source_name = ws.cell(r, 2).value
        rows.append({
            "row": r,
            "source_name": source_name,
            "zh_cle": source_name,
            "zh_joyoland": canonical.get("zh_cn") if canonical else None,
            "canonical_reference": canonical,
            "source_kind": ws.cell(r, 1).value,
            "source_caster": ws.cell(r, 3).value,
            "match_status": "row_order",
            "name_equals_canonical": bool(canonical and canonical.get("zh_cn") == source_name),
        })
    return {
        "sheet": ws.title,
        "kind": "reference_skill_source_variants",
        "row_count": len(rows),
        "rows": rows,
    }


def build_row_order_items(ws, canonical_items: list[dict], kind: str) -> dict:
    rows = []
    for idx, r in enumerate(chunk_rows(ws)):
        canonical = canonical_items[idx] if idx < len(canonical_items) else None
        source_name = ws.cell(r, 1).value
        rows.append({
            "row": r,
            "source_name": source_name,
            "zh_cle": source_name,
            "zh_joyoland": canonical.get("zh_cn") if canonical else None,
            "canonical_item": as_variant(canonical),
            "match_status": "exact_name" if canonical and canonical.get("zh_cn") == source_name else "row_order",
            "name_equals_canonical": bool(canonical and canonical.get("zh_cn") == source_name),
        })
    return {
        "sheet": ws.title,
        "kind": kind,
        "row_count": len(rows),
        "rows": rows,
    }


def build_core_quartz(ws, items_by_name: dict[str, list[dict]]) -> dict:
    groups = []
    current_name = None
    current_rows = []

    def flush() -> None:
        nonlocal current_name, current_rows
        if current_name is None:
            return
        canonical_name = CORE_ALIAS.get(current_name)
        canonical_candidates = items_by_name.get(canonical_name, []) if canonical_name else []
        groups.append({
            "source_name": current_name,
            "zh_cle": current_name,
            "zh_joyoland": canonical_name,
            "canonical_candidates": [as_variant(row) for row in canonical_candidates],
            "match_status": "manual_alias_ambiguous" if len(canonical_candidates) > 1 else ("manual_alias" if canonical_candidates else "unmatched"),
            "rows": current_rows,
        })
        current_name = None
        current_rows = []

    for r in chunk_rows(ws):
        source_name = repair_core_text(ws.cell(r, 1).value)
        if source_name != current_name:
            flush()
            current_name = source_name
        current_rows.append({
            "row": r,
            "level": ws.cell(r, 4).value,
            "attributes": {
                "hp": ws.cell(r, 5).value,
                "ep": ws.cell(r, 6).value,
                "str": ws.cell(r, 7).value,
                "def": ws.cell(r, 8).value,
                "ats": ws.cell(r, 9).value,
                "adf": ws.cell(r, 10).value,
                "spd": ws.cell(r, 11).value,
            },
            "core_magic": repair_core_text(ws.cell(r, 3).value),
            "elements": repair_core_text(ws.cell(r, 12).value),
            "effect": repair_core_text(ws.cell(r, 13).value),
            "obtain": repair_core_text(ws.cell(r, 14).value),
        })
    flush()
    return {
        "sheet": ws.title,
        "kind": "core_quartz_source_variants",
        "group_count": len(groups),
        "groups": groups,
    }


def main() -> None:
    glossary = load_json(CANONICAL)
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    items = glossary.get("items", [])
    items_by_name = index_items(items)

    prop_items = [row for row in items if row.get("category_zh") == "道具-普通"]
    equip_items = [row for row in items if row.get("category_zh", "").startswith("装备-")]
    circuit_items = [row for row in items if row.get("category_zh") == "回路-普通"]

    result = {
        "meta": {
            "source_workbook": str(WORKBOOK),
            "canonical_glossary": str(CANONICAL),
            "layer_model": ["canonical", "zh_joyoland", "zh_cle"],
            "notes": [
                "The workbook is treated as the CLE source layer.",
                "The existing ao_zh_glossary.json zh_cn values are treated as the Joyoland source layer.",
                "Canonical IDs and save-editor metadata remain unchanged.",
                "Core quartz uses a manual alias map because CLE and Joyoland diverge on several names.",
            ],
        },
        "sheets": {
            "人物经验值": build_characters(wb["人物经验值"], glossary),
            "魔法技能": build_reference_skills(wb["魔法技能"], glossary),
            "道具": build_row_order_items(wb["道具"], prop_items, "item_source_variants"),
            "装备": build_row_order_items(wb["装备"], equip_items, "equipment_source_variants"),
            "核心回路": build_core_quartz(wb["核心回路"], items_by_name),
            "工作表6": build_row_order_items(wb["工作表6"], circuit_items, "normal_quartz_source_variants"),
        },
    }

    payload = json.dumps(result, ensure_ascii=False, indent=2)
    OUT_WORKSPACE.write_text(payload, encoding="utf-8")
    OUT_PUBLISH.write_text(payload, encoding="utf-8")

    print(OUT_WORKSPACE)
    print(OUT_PUBLISH)
    print(json.dumps({
        "人物经验值": result["sheets"]["人物经验值"]["row_count"],
        "魔法技能": result["sheets"]["魔法技能"]["row_count"],
        "道具": result["sheets"]["道具"]["row_count"],
        "装备": result["sheets"]["装备"]["row_count"],
        "核心回路": result["sheets"]["核心回路"]["group_count"],
        "工作表6": result["sheets"]["工作表6"]["row_count"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
