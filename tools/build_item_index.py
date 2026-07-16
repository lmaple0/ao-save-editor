from __future__ import annotations

from pathlib import Path
import json

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
PUBLISH_ROOT = ROOT.parent / "publish" / "package"
WORKBOOK = Path.home() / "Downloads" / "ao_data_cn.xlsx"
GLOSSARY_JSON = ROOT / "ao_zh_glossary.json"
OUT_JSON = ROOT / "ao_item_index.json"
OUT_JSON_PUBLISH = PUBLISH_ROOT / "ao_item_index.json"

WEAPON_CATEGORY_ORDER = [
    "equipment_weapon_Lloyd",
    "equipment_weapon_Elie",
    "equipment_weapon_Tio",
    "equipment_weapon_Randy",
    "equipment_weapon_Noel",
    "equipment_weapon_Lazy",
    "equipment_weapon_Rixia",
    "equipment_weapon_Dudley",
    "equipment_weapon_Arios",
    "equipment_weapon_Zeit",
]

NORMAL_QUARTZ_ALIASES = {
    "耀脈": "耀脉",
    "真心": "丹精",
    "捉弄": "恶戏",
    "练气": "炼气",
    "黑暗之刃": "暗之刃",
    "黑暗之刃2": "暗之刃2",
    "破足之牙": "破脚之牙",
    "异香": "美臭",
    "妨碍1": "妨害1",
    "妨碍2": "妨害2",
    "鹰眼": "鹰目",
}

CORE_QUARTZ_ALIASES = {
    "护盾": "盾牌",
    "神盾": "神佑",
    "小丑": "小丑",
    "圣典": "法则",
    "权杖": "王权",
    "智者": "贤者",
    "力量": "力量",
    "战斧": "战斧",
    "骑兵": "骑士",
    "妖精": "妖精",
    "羽翼": "羽翼",
    "草薙": "草薤",
    "利爪": "利爪",
    "坚守": "守护",
    "罪恶": "罪恶",
    "白金": "白金",
    "永恒": "无限",
    "纹章": "纹章",
    "幻象": "幻象",
    "魅影": "魅影",
    "灵猫": "灵猫",
}


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def rows_from_sheet(workbook, sheet_name: str, column: int = 1) -> list[str]:
    ws = workbook[sheet_name]
    return [
        ws.cell(row, column).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, column).value
    ]


def index_by_joyoland(items: list[dict]) -> dict[str, list[dict]]:
    result: dict[str, list[dict]] = {}
    for item in items:
        name = item.get("zh_cn")
        if name:
            result.setdefault(name, []).append(item)
    return result


def assign_exact_for_categories(
    names: dict[int, str],
    source_names: list[str],
    by_name: dict[str, list[dict]],
    categories: set[str],
) -> None:
    for source_name in source_names:
        matches = [item for item in by_name.get(source_name, []) if item.get("category") in categories]
        if len(matches) == 1:
            names[int(matches[0]["id_dec"])] = source_name


def assign_row_order(names: dict[int, str], source_names: list[str], canonical_items: list[dict]) -> None:
    for source_name, item in zip(source_names, canonical_items):
        names[int(item["id_dec"])] = source_name


def collect_cle_names(glossary_items: list[dict]) -> dict[int, str]:
    names: dict[int, str] = {}
    by_name = index_by_joyoland(glossary_items)
    if not WORKBOOK.exists():
        return names

    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    item_names = rows_from_sheet(workbook, "道具")
    equipment_names = rows_from_sheet(workbook, "装备")
    quartz_names = rows_from_sheet(workbook, "工作表6")

    assign_exact_for_categories(names, item_names, by_name, {"props_normal", "props_cooking"})
    assign_exact_for_categories(names, quartz_names, by_name, {"circuit_normal"})

    props_normal = [item for item in glossary_items if item.get("category") == "props_normal"]
    assign_row_order(names, item_names[:len(props_normal)], props_normal)

    weapon_items = [
        item
        for category in WEAPON_CATEGORY_ORDER
        for item in glossary_items
        if item.get("category") == category
    ]
    assign_row_order(names, equipment_names[:len(weapon_items)], weapon_items)

    for source_name, joyoland_name in NORMAL_QUARTZ_ALIASES.items():
        matches = by_name.get(joyoland_name, [])
        if len(matches) == 1:
            names[int(matches[0]["id_dec"])] = source_name

    for source_name, joyoland_name in CORE_QUARTZ_ALIASES.items():
        for item in by_name.get(joyoland_name, []):
            if item.get("category") == "circuit_core":
                names[int(item["id_dec"])] = source_name

    return names


def build_index() -> dict:
    glossary = load_json(GLOSSARY_JSON)
    glossary_items = glossary.get("items", [])
    cle_names = collect_cle_names(glossary_items)
    existing = load_json(OUT_JSON) if OUT_JSON.exists() else {"items": []}
    existing_cle = {
        int(row["id_dec"]): row.get("zh_cle")
        for row in existing.get("items", [])
        if row.get("zh_cle")
    }

    items = []
    for row in sorted(glossary_items, key=lambda item: int(item["id_dec"])):
        item_id = int(row["id_dec"])
        items.append({
            "id_dec": item_id,
            "id_hex": row.get("id_hex"),
            "category": row.get("category"),
            "category_zh": row.get("category_zh"),
            "zh_joyoland": row.get("zh_cn"),
            "zh_cle": existing_cle.get(item_id) or cle_names.get(item_id),
            "en": row.get("en"),
            "ja": row.get("ja"),
        })

    return {
        "meta": {
            "source": [str(GLOSSARY_JSON), str(WORKBOOK)],
            "notes": [
                "Runtime item index for the save editor.",
                "zh_joyoland comes from ao_zh_glossary.json and is keyed by ID from Ouroboros/Falcom ItemNameMap.py.",
                "Joyoland source: https://github.com/Ouroboros/Falcom/blob/master/ED7/Decompiler/GameData/ItemNameMap.py",
                "zh_cle is intentionally conservative: props, weapons, normal quartz aliases, and core quartz aliases only.",
            ],
        },
        "items": items,
    }


def main() -> None:
    data = build_index()
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    OUT_JSON.write_text(payload, encoding="utf-8")
    OUT_JSON_PUBLISH.write_text(payload, encoding="utf-8")
    print(OUT_JSON)
    print(OUT_JSON_PUBLISH)
    print(json.dumps({"items": len(data["items"])}, ensure_ascii=False))


if __name__ == "__main__":
    main()
